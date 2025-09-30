#!/usr/bin/python3
import os, sys
import re
import logging
import pandas as pd
import json, json5
from openpyxl.styles import PatternFill, Font
from openpyxl import Workbook
import argparse

from icn3d_igstrand_refnum import get_igstrand_reference, check_filename_exist
from igstrand_domain_mapping import get_igmap_domain


logging.basicConfig(
    filename='igstrand.log',
    filemode = 'w',
    level=logging.INFO,
    format='%(asctime)s.%(msecs)03d %(levelname)s: %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
)

def split_string(number_string):
    """
    This will split the number into strand number and number part
    input: "A'1248a"
    ouput: A', 1248a
    """
    match = re.search(r'^([^0-9]*)([0-9].*)$', number_string)
    return match.groups() if match else None


def sort_dict_by_key(dictionary):
    """ 
    Here's a function that sorts a dictionary by its keys based on the following rules:
    If the key is an integer, sort by the integer value.
    If the key is an integer followed by a single alphabet letter, 
    sort by the integer value first, then sort by the alphabet letter.

    """
    sorted_keys = sorted(dictionary.keys(), key=lambda k: tuple(int(x) if x.isdigit() else x for x in re.findall(r'\d+|\D+', k)))

    return {k: dictionary[k] for k in sorted_keys}


def get_igmap_info(pdb_chain_domain_input, file_path):
    """
    This will get the information based on the pdb_chain_domain.
    This will check the files whether exits or not then will download.
    """
    if get_igstrand_reference(pdb_chain_domain_input[0],  file_path +"number_mapping_files/"): #
        map_igstrand_info = get_igmap_domain(pdb_chain_domain_input, "igstrand", file_path+"number_mapping_files/")
        map_igref_key = f"{pdb_chain_domain_input[0]}_{pdb_chain_domain_input[1]}_{pdb_chain_domain_input[2]}" 
        if not map_igstrand_info:
            parse_ig_refdata = {map_igref_key: {'3Ddomain_order': "",'3dD_res_range':"", 'igD_res_range':"", 'refpdbname':"", 'tmscore':"", 
     'seqid':"", 'nresAlign':"", 'Igtype': "",'undefined_info':[],"igstrand_data":{}}}
           
        else:
            parse_ig_refdata = {map_igref_key:map_igstrand_info}

        return parse_ig_refdata


def get_all_igrefnum_keys(all_ig_data):
    """
    This will create all possible keys for given domain to create a
    alignment.
    """
    all_keys = {re.sub('[a-z]', '', y.split("_")[0]) for i in all_ig_data for x in i for y in i[x]['igstrand_data']if i[x]['igstrand_data']}

    all_mapping_value = {split_string(key)[1]:split_string(key)[0] for key in all_keys}
     # # sort by key to maintain the sequence
    sorted_mapping_value = sort_dict_by_key(all_mapping_value)
    sorted_mapping_value = {value+key:[] for key, value in sorted_mapping_value.items()}

    return sorted_mapping_value


def fill_excel_cell_residues(ig_data, ignumkey, out_cell, color_map):
    """
    ig_data: mapping key and value dictionary
    ignumkey: C'4548
    out_cell: destination cell
    color_map: color map dictionary for strands

    """
    if ignumkey in ig_data:
        res_id, loop_assign = ig_data[ignumkey]
        out_cell.value = res_id
        # if 50 number
        if key_look[-2:] == "50":
            hex_code = "FFD700" # this for last 50 residues
            font = Font(size=16)
            out_cell.font = font
            fill = PatternFill(start_color=hex_code, end_color=hex_code, fill_type='solid')
            out_cell.fill = fill
            out_cell.font = font1
        # now check loop.
        elif loop_assign:
            hex_code = color_map["loop"]
            #print(hex_code)
            fill = PatternFill(start_color=hex_code, end_color=hex_code, fill_type='solid')
            out_cell.fill = fill
            out_cell.font = font1
        else:
            #rest
            # find the strand

            pure_strand = split_string(key_look)[0].strip("+-_") # this is for color purpose.
            hex_code = color_map.get(pure_strand, 'FFFFFF')
            fill = PatternFill(start_color=hex_code, end_color=hex_code, fill_type='solid')
            out_cell.fill = fill
            out_cell.font = font1



    else:

        out_cell.value = ""


    return

def fill_excel_reference_info(sorted_all_igrefnum, ref_headers, ws1, wb1, font_size):
    """
    This will fill the excel sheet based on the reference information.
    """
    # fill the column header
    for col, header in enumerate(ref_headers, start=1): #excel col starts from 1
        ws1.cell(row=1, column=col, value= header)
        font = Font(size=font_size+2)  #write header two font larger                                                    
        ws1.cell(row=1, column=col).font = font

    # now fill the residues ig refnum headers. 
    for col, hd in enumerate(sorted_all_igrefnum):
        # filled header with number
        destination_header_cell = ws1.cell(row = 1, column = col+ len(ref_headers) + 1) # first column for header
        destination_header_cell.value = hd 
        font = Font(size= font_size+2)
        destination_header_cell.font = font


    return


def process_all_excel_cell(ws1, wb1, all_igfile_info, sorted_map, ref_headers, font_size):
    """
    """
    # print(all_igfile_info)
    
  
    for row_val, file in enumerate(all_igfile_info):
        for stru in file:
            font1 = Font(size=font_size)
            ws1.cell(row = row_val+2, column = 1, value = stru)
            igstrand_data = file[stru]["igstrand_data"]

            headers_not_str_undefined = [elem for elem in ref_headers if elem not in ["structure", "undefined_info"]]
            for col, header in enumerate(headers_not_str_undefined, start=2):
                ws1.cell(row=row_val+2, column=col, value= file[stru][header])
                ws1.font = font1

            if isinstance(file[stru]["undefined_info"], list):
                 # Convert empty list to empty string
                ws1.cell(row = row_val+2, column = len(ref_headers), value = "")
                ws1.font = font1
            else:
                ws1.cell(row = row_val+2, column = len(ref_headers), value = file[stru]["undefined_info"])
                ws1.font = font1

            for col, key_look in enumerate(sorted_map):
                destination_cell = ws1.cell(row = row_val+2, column = col+ len(ref_headers) + 1)
               
                fill_excel_cell_residues(igstrand_data, key_look, destination_cell, color_dict, font_size)

    return ws1, wb1


# put all code together.


     



if __name__== "__main__":

    # input_file_path = os.getenv('input_file_path')
    # output_file_path = os.getenv('output_file_path')
    # node_js_file_path = os.getenv('node_js_file_path')
    # numbering_name = os.getenv('numbering_name')
        
    input_file_path = "../input/"
    output_file_path = "../output/"
    node_js_file_path = "node_js_script/"
    numbering_name = "igstrand"

    # deal the color

  
    color_dict = {"A": "9400D3", "A'": "9400D3", "B": "ba55d3", "C": "0000FF", "C'": "6495ED",
              "C''": "006400", "D": "00FF00", "E": "FFD700", "F": "FF8C00", "G": "FF0000",
              "loop": "CCCCCC"}

    headers = ['structure', '3dD_res_range', 'igD_res_range', 'refpdbname', 'tmscore', 
    'seqid', 'nresAlign', 'undefined_info']

    # input_file_data = [("7LQ7", "H", 1), 
    #                 ("1CD8", "A", "1"), 
    #                 ("7N4I", "L", "1"),
    #                 ("1RHH", "B", "1"),
    #                 ("5ESV", "A", "1"),
    #                 ("5ESV", "B", "1")]
    
    wb = Workbook()
    ws = wb.active
    all_file_info = []

    print(f"Starting 1D alignment..")

    for pdb_chain_domain in input_file_data:
        
        all_file_info.append(get_igmap_info(pdb_chain_domain, input_file_path))
            
    sorted_all_mapping_value = get_all_igrefnum_keys(all_file_info)
    #write column headers

    fill_excel_reference_info(sorted_all_mapping_value, headers, ws, wb, font_size=12)

    ws, wb = process_all_excel_cell(ws, wb, all_file_info, sorted_all_mapping_value, headers, font_size=12)



    wb.save(f"{output_file_path}1D_mapping_{numbering_name.lower()}.xlsx")
print(f"A 1D alignment file, 1D_mapping_{numbering_name.lower()}.xlsx, is created in the {output_file_path}")
print()



            


    # 