#!/usr/bin/python3

import re, os, json
from typing import Optional, Tuple, Dict, List
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Border,  PatternFill, Alignment
from copy import copy
import argparse



from igstrand_domain_mapping import get_igmap_domain
from icn3d_igstrand_refnum import get_igstrand_reference, check_filename_exist

def split_string(number_string):
    """
    This will split the number into strand number and number part
    input: "A'1248a"
    ouput: A', 1248a
    """
    match = re.search(r'^([^0-9]*)([0-9].*)$', number_string)
    return match.groups() if match else None

def open_template_file(ig_type: str, numbering_name: str, file_path: str) -> Tuple[Worksheet, Workbook]:
    """
    Open the corresponding template file based on the provided parameters.

    Args:
    - ig_type (str): The type of Ig.
    - numbering_name (str): The numbering name.
    - file_path (str): The path to the template files.

    Returns:
    - Tuple[Worksheet, Workbook]: A tuple containing the opened worksheet and workbook.
    """
    try:
        wb = load_workbook(f"{file_path}{numbering_name.lower()}_template_{ig_type}.xlsx")
        ws = wb.active
    except FileNotFoundError as e:
        raise FileNotFoundError(f"Template file is not found: {e}")

    return ws, wb




def modify_excel_residue_mapping(id_chain_domain: Tuple[str, str, str], map_res: Dict[str, Tuple[str, str]],ref_struct: str, ig_type: str, ws: Worksheet, ws_out: Worksheet, template_length: Tuple[int, int], num_columns: int, color_dict: Dict[str, str]) -> Worksheet:
    """
    Fill the residues information in the Excel worksheet based on the provided mapping and write the modified values to a new Excel file with shifted columns.

    Args:
    - id_chain_domain (Tuple[str, str, str]): A tuple containing id, chain, and domain information.
    - map_res (Dict[str, Tuple[str, bool]]): A dictionary containing residue mapping information.
    - ref_struct (str): reference structure mapping name.
    - ig_type (str): The type of Ig.
    - ws (Worksheet): The Excel worksheet to modify.
    - ws_out: Worksheet: destination worksheet to write.
    - template_range (tuple[int, int]):
    - num_columns (int): The number of columns to shift.
    - output_file_path (str): The path to save the modified Excel file.

    Returns:
    - Worksheet
    # """
    # row_range = template_range["row_range"] # may be we can do also automation.
    # col_range = template_range[f"{ig_type}_column_range"]

    # Shift columns to the right and fill the new columns
    for row in range(1, template_length[0] + 1):
        for col in range(template_length[1], 0, -1):
            source_cell = ws.cell(row=row, column=col)
       
            new_column = col + num_columns
            destination_cell = ws_out.cell(row=row, column=new_column)
            cell_value = source_cell.value
            # color code based on strand not on template so
            # if len(cell_value)== 4 and cell_value[0] in [str(i) for i in range(10)]: # this means this is igstrand number.
            

            res_id, loop_assign = map_res.get(str(cell_value), ("", False))
            if res_id:
                if loop_assign:
                    color_code = color_dict["loop"]
                else: # residues exits.
                    if str(cell_value)[-2:] == "50":
                        color_code= "FFFF00"
                    else: #number exits but not anchor
                        color_code = color_dict[str(cell_value)[0]] # based on first digit #you can copy color also from template. 
                destination_cell.value = res_id
                destination_cell.fill = PatternFill(start_color= color_code, end_color=color_code, fill_type='solid')

            else:
                destination_cell.fill  = copy(source_cell.fill)

                if cell_value  == ig_type:
                    destination_cell.value =ig_type +"_" +"_".join([str(elem) for elem in id_chain_domain])
                    destination_cell_next = ws_out.cell(row=row, column=new_column+3) # write in next 6th column
                    destination_cell_next.value = ref_struct
                    destination_cell_next.font = Font(bold=True, size=16)
 
                    
                    
                elif cell_value and len(str(cell_value))== 4 and str(cell_value)[0] in [str(i) for i in range(10)]: 
                    destination_cell.value = ""

                else:
                    destination_cell.value = cell_value
                    
            destination_cell.alignment = Alignment(horizontal='center')
            destination_cell.font = copy(source_cell.font)
            destination_cell.border = copy(source_cell.border)
            

    return ws_out


def read_input_file(file_path: str) -> List[Tuple[str, str, str]]:
    """
    Read the input file containing pdb, chain, and domain information.

    Args:
    - file_path (str): The path to the input file.

    Returns:
    - List[Tuple[str, str, str]]: A list of tuples containing pdb, chain, and domain information.
    """

   
    input_file = list()

    with open(file_path) as file_input:
        for file_lines in file_input:
            if file_lines.strip():
                field = file_lines.strip().split(" ") 
                if len(field) == 3:
                    input_file.append((field[0].upper(), field[1], field[2]))

                else:
                    print(f"Make sure {field} has three value: pdbid chain domain")

    return input_file




if __name__== "__main__":
    parser = argparse.ArgumentParser(description='Process input file')
    parser.add_argument('-f', '--file', help='Input file name', required=True)
    args = parser.parse_args()

    # input_file_path= "../input/"
    # output_file_path = "../output/"
    # node_js_file_path = "node_js_script/"
    # numbering_name = "igstrand"
    # template_row_col = {
    #     "V_column_range": 21,
    #     "V_row_range" : 47}


    input_file_path = os.getenv('input_file_path')
    output_file_path = os.getenv('output_file_path')
    node_js_file_path = os.getenv('node_js_file_path')
    numbering_name = os.getenv('numbering_name')

    ##Access the dictionary from the environment variable
    template_row_col = json.loads(os.environ['template_row_col'])


    input_file_data  = read_input_file(args.file)
    output_save_name = args.file.split(".")[0]
    #input_file_data  = read_input_file("test.txt")
    #input_file_data  = read_input_file("input_IgC2_template.txt")

    print(f"Starting 2D alignment..")


    color_dict = {"1": "9400D3",  "2": "ba55d3", "3": "0000FF", "4": "6495ED",
                  "5": "006400", "6": "00FF00", "7": "FFD700", "8": "FF8C00", "9": "FF0000",
                  "loop": "CCCCCC"}


    # Create a new workbook for output
    wb_out = Workbook()
    ws_out= wb_out.active

    num_input = len(input_file_data)
    
    columns_to_shift = 0 # start from first column

    # put template in one 

    for pdb_chain_domain in input_file_data:
        # check if file exits otherwise call node js to download.
        if get_igstrand_reference(pdb_chain_domain[0], input_file_path +"number_mapping_files/"): #

            map_igstrand_info = get_igmap_domain(pdb_chain_domain, "igstrand", input_file_path+"number_mapping_files/")

            if map_igstrand_info is not None:
                map_res_ig = map_igstrand_info.get("igstrand_data")
                #map_res id has strand ids aslo so remove the
                map_res_nostrand_letter = {split_string(x)[1]:y for x,y in map_res_ig.items()}
                ig_match_type = map_igstrand_info.get("Igtype")
                # check if IgV type
                if ig_match_type == "IgV":
                    map_res_strand_letter = set(split_string(x)[0] for x in map_res_ig)
                    if "A" in map_res_strand_letter and "A'" in map_res_strand_letter:
                        ig_match_type_template = "IgV_A_Adash"
                    elif "A'" in map_res_strand_letter:
                        ig_match_type_template = "IgV_Adash"
                    elif "A" in map_res_strand_letter:
                        ig_match_type_template = "IgV_A"
                else:
                    ig_match_type_template = ig_match_type

                map_ref_pdb = map_igstrand_info.get('refpdbname')
                # here for template not found
                try:
                    ws, wb = open_template_file(ig_match_type_template, numbering_name, input_file_path+"/igstrand_template/") # open template
                except FileNotFoundError as e:
                    print(f"Skipping {pdb_chain_domain}:{e}")
                    print()
                    continue


                # if given select from template_row_col else use  V_row_range and V_column_range



                template_row_col_length =  (template_row_col.get(f"{ig_match_type}_row_range", 
                    template_row_col["V_row_range"]), template_row_col.get(f"{ig_match_type}_column_range", template_row_col["V_column_range"]))


                ws_out = modify_excel_residue_mapping(pdb_chain_domain, map_res_nostrand_letter, map_ref_pdb, ig_match_type, ws, ws_out, template_row_col_length, columns_to_shift, color_dict)

                # assume column length of each template is


                columns_to_shift += template_row_col_length[1]  # width of template column
                wb_out.save(f"{output_file_path}2D_mapping_{numbering_name.lower()}.xlsx")
                
               
            else:
                print(f"Ig domain is not found in  {pdb_chain_domain}. 2D figure is not created.")
                print()
             
        
print(f"2D figures are created in {output_file_path}2D_mapping_{output_save_name}{numbering_name.lower()}.xlsx.")
print()    

   
    
    
    

